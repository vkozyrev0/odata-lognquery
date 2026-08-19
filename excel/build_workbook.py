"""Generate OData-Query-Scenarios.xlsx from the scenario docs and .pq scripts."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent
NAVY = "1E3A5F"
TEAL = "0F766E"
SLATE = "334155"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
SECTION_FILL = PatternFill("solid", fgColor="E2E8F0")
ALT_FILL = PatternFill("solid", fgColor="F8FAFC")
PRO_FILL = PatternFill("solid", fgColor="ECFDF5")
CON_FILL = PatternFill("solid", fgColor="FEF2F2")
WHITE = Font(name="Calibri", color="FFFFFF", bold=True, size=11)
TITLE = Font(name="Calibri", bold=True, size=18, color=NAVY)
SUBTITLE = Font(name="Calibri", bold=True, size=13, color=TEAL)
BODY = Font(name="Calibri", size=11, color=SLATE)
BOLD = Font(name="Calibri", bold=True, size=11, color=SLATE)
CODE = Font(name="Consolas", size=9, color="0F172A")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)

PRODUCTS = [
    (1, "Chai", "Beverages", 4.50, 0, True),
    (2, "Chang", "Condiments", 7.75, 7, False),
    (3, "Aniseed Syrup", "Produce", 11.00, 14, False),
    (4, "Chef Anton's Cajun Seasoning", "Seafood", 14.25, 21, False),
    (5, "Grandma's Boysenberry Spread", "Dairy", 17.50, 28, False),
    (6, "Uncle Bob's Organic Dried Pears", "Beverages", 20.75, 35, False),
    (7, "Northwoods Cranberry Sauce", "Condiments", 24.00, 42, False),
    (8, "Mishi Kobe Niku", "Produce", 27.25, 49, True),
    (9, "Ikura", "Seafood", 30.50, 56, False),
    (10, "Queso Cabrales", "Dairy", 33.75, 63, False),
]


def set_col_widths(ws: Worksheet, widths: dict[int, float]) -> None:
    for index, width in widths.items():
        ws.column_dimensions[get_column_letter(index)].width = width


def header_row(ws: Worksheet, row: int, values: list[str]) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row, col, value)
        cell.fill = HEADER_FILL
        cell.font = WHITE
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.row_dimensions[row].height = 22


def write_wrapped(ws: Worksheet, row: int, col: int, value: str, fill: PatternFill | None = None) -> None:
    cell = ws.cell(row, col, value)
    cell.font = BODY
    cell.alignment = WRAP
    cell.border = THIN
    if fill:
        cell.fill = fill


def add_m_sheet(wb: Workbook, title: str, path: Path) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = TEAL
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells("A1:B1")
    ws["A2"] = (
        "Paste this M into Excel: Data → Get Data → From Other Sources → Blank query → Advanced Editor. "
        "The sample service must be running at http://127.0.0.1:5268."
    )
    ws["A2"].font = BODY
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:B2")
    ws.row_dimensions[2].height = 36
    ws["A4"] = "Power Query M"
    ws["A4"].font = SUBTITLE
    text = path.read_text(encoding="utf-8")
    ws["A5"] = text
    ws["A5"].font = CODE
    ws["A5"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A5:B5")
    ws.row_dimensions[5].height = min(420, 18 + text.count("\n") * 14)
    set_col_widths(ws, {1: 120, 2: 20})
    ws.freeze_panes = "A5"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False


def add_preview_sheet(wb: Workbook, title: str, caption: str) -> None:
    ws = wb.create_sheet(title)
    ws["A1"] = title
    ws["A1"].font = TITLE
    ws.merge_cells("A1:F1")
    ws["A2"] = caption
    ws["A2"].font = BODY
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 36
    header_row(ws, 4, ["Id", "Name", "Category", "Price", "UnitsInStock", "Discontinued"])
    for offset, row in enumerate(PRODUCTS):
        excel_row = 5 + offset
        for col, value in enumerate(row, start=1):
            cell = ws.cell(excel_row, col, value)
            cell.font = BODY
            cell.border = THIN
            cell.alignment = Alignment(vertical="center")
            if offset % 2:
                cell.fill = ALT_FILL
            if col == 4:
                cell.number_format = "0.00"
    set_col_widths(ws, {1: 8, 2: 36, 3: 16, 4: 12, 5: 16, 6: 16})
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:F14"


def build() -> Path:
    wb = Workbook()

    cover = wb.active
    cover.title = "Overview"
    cover.sheet_properties.tabColor = NAVY
    cover["A1"] = "OData long-running query scenarios"
    cover["A1"].font = TITLE
    cover.merge_cells("A1:B1")
    cover["A3"] = (
        "This workbook documents the three first-request strategies in the Angular demo "
        "and how they behave when OData paging is used on large sets. "
        "The sample catalog is 5000 products with a server page size of 500 "
        "(@odata.nextLink until the last page). "
        "The Power Query M on the later sheets uses the same page-following algorithm as the web app."
    )
    cover["A3"].font = BODY
    cover["A3"].alignment = WRAP
    cover.merge_cells("A3:B3")
    cover.row_dimensions[3].height = 52

    cover["A5"] = "How to use"
    cover["A5"].font = SUBTITLE
    steps = [
        "1. Start the sample API:  dotnet run --project ODataLongQuery",
        "2. Confirm http://127.0.0.1:5268/odata/Products in a browser.",
        "3. In Excel 365 (Windows): Data → Get Data → From Other Sources → Blank query.",
        "4. Open Advanced Editor and paste one of the M sheets (or the .pq files in excel/power-query/).",
        "5. If Excel asks about privacy, choose Ignore / Anonymous for localhost.",
        "6. Refresh. Excel still waits until M finishes; async only avoids HTTP/gateway timeouts.",
    ]
    for index, line in enumerate(steps):
        cover.cell(6 + index, 1, line).font = BODY

    cover["A13"] = "Sheets"
    cover["A13"].font = SUBTITLE
    sheets_info = [
        ("Comparison", "Pros and cons of sync, async, and async with wait."),
        ("Paging at scale", "How @odata.nextLink interacts with each scenario."),
        ("M — Sync", "Web.Contents loop: no Prefer, follow @odata.nextLink (5000 / 500)."),
        ("M — Async", "Same loop; Prefer: respond-async + poll on every page."),
        ("M — Async wait", "Same loop; wait=2 on each page (4s wait-mode delay → 202)."),
        ("Preview — Sync", "Sample Products table as if loaded synchronously."),
        ("Preview — Async", "Same preview; first request would have been 202 + poll."),
        ("Preview — Wait", "Same preview; first request may have returned 200 within wait."),
    ]
    header_row(cover, 14, ["Sheet", "Contents"])
    for offset, (name, desc) in enumerate(sheets_info):
        write_wrapped(cover, 15 + offset, 1, name)
        write_wrapped(cover, 15 + offset, 2, desc, ALT_FILL if offset % 2 else None)
        cover.row_dimensions[15 + offset].height = 28
    set_col_widths(cover, {1: 28, 2: 92})
    cover.freeze_panes = "A15"
    cover.sheet_view.showGridLines = False

    comparison = wb.create_sheet("Comparison")
    comparison.sheet_properties.tabColor = TEAL
    comparison["A1"] = "Pros and cons"
    comparison["A1"].font = TITLE
    comparison.merge_cells("A1:D1")
    comparison["A2"] = (
        "All three modes return the same Products entity set. They differ only in whether "
        "the original HTTP request is held open. See docs/query-scenarios.md for the full write-up."
    )
    comparison["A2"].font = BODY
    comparison["A2"].alignment = WRAP
    comparison.merge_cells("A2:D2")
    comparison.row_dimensions[2].height = 40

    header_row(
        comparison,
        4,
        [
            "Topic",
            "1. Synchronous",
            "2. Async (respond-async)",
            "3. Async with wait",
        ],
    )
    rows = [
        (
            "HTTP on the first request",
            "GET held until 200. No Prefer header.",
            "GET returns 202 immediately with Location /async/{id}.",
            "GET held up to N seconds; 200 if done, else 202.",
        ),
        (
            "Demo UI radio",
            "Synchronous — wait on this request",
            "Async — Prefer: respond-async, then poll",
            "Async with wait — hold up to N seconds",
        ),
        (
            "Power Query",
            "Web.Contents, no Prefer; follow nextLink (same as Angular sync).",
            "Web.Contents + respond-async + poll on every page, then nextLink.",
            "Same loop; wait=N on page 1, wait=min(N,5) on later pages.",
        ),
        (
            "Pros",
            "One round trip per page; easy errors; best for cheap pages.",
            "Avoids gateway timeouts; cancel via DELETE; service can drop waiters.",
            "Fast pages stay one hop; slow pages fall back to 202; less poll chatter.",
        ),
        (
            "Cons",
            "Timeouts; blocked UI; no job to cancel; held connections.",
            "Job store, poll QPS, PQ recursion; jobs die on recycle in this sample.",
            "N must sit under the shortest timeout; wrong N becomes sync or always-202.",
        ),
        (
            "Cancel",
            "Not in the protocol.",
            "DELETE /async/{id} while running.",
            "DELETE only if the service returned 202.",
        ),
        (
            "When to use",
            "Each page finishes well inside every timeout.",
            "First page (or whole query) can exceed timeouts.",
            "Mixed pages: most cheap, outliers slow. One client path.",
        ),
    ]
    for offset, values in enumerate(rows):
        excel_row = 5 + offset
        for col, value in enumerate(values, start=1):
            fill = None
            if values[0] == "Pros":
                fill = PRO_FILL
            elif values[0] == "Cons":
                fill = CON_FILL
            elif offset % 2:
                fill = ALT_FILL
            write_wrapped(comparison, excel_row, col, value, fill)
        comparison.row_dimensions[excel_row].height = 64
    set_col_widths(comparison, {1: 22, 2: 36, 3: 40, 4: 42})
    comparison.freeze_panes = "A5"
    comparison.auto_filter.ref = "A4:D11"
    comparison.sheet_view.showGridLines = False

    paging = wb.create_sheet("Paging at scale")
    paging.sheet_properties.tabColor = "B45309"
    paging["A1"] = "Very large sets and OData pagination"
    paging["A1"].font = TITLE
    paging.merge_cells("A1:D1")
    paging["A2"] = (
        "Paging (@odata.nextLink or $top/$skip) splits a result into pages. "
        "HTTP 202 splits wait time off the original connection. A large query can need both. "
        "202 applies to one HTTP resource — usually the first page. Each nextLink is a new request."
    )
    paging["A2"].font = BODY
    paging["A2"].alignment = WRAP
    paging.merge_cells("A2:D2")
    paging.row_dimensions[2].height = 48

    header_row(
        paging,
        4,
        [
            "Concern",
            "Sync",
            "Async",
            "Async with wait",
        ],
    )
    paging_rows = [
        (
            "Who follows nextLink in Excel?",
            "This sample’s M loops nextLink (OData.Feed would also).",
            "M loops nextLink after the monitor returns 200 (new job per page).",
            "Same as async; first response may already be 200; later wait is 5s.",
        ),
        (
            "First page is a heavy scan",
            "Held GET; gateway timeout risk is highest here.",
            "202 + poll until the first page exists.",
            "Wait=N on page 1 (e.g. 20–60s, still under the gateway limit).",
        ),
        (
            "Later pages are cheap seeks",
            "Still blocking, but usually fine if each page is fast.",
            "Do not start a new job per page unless needed; prefer sync GET on nextLink.",
            "Use a smaller wait on continuation (NextPageWaitSeconds in the M).",
        ),
        (
            "Later pages are also heavy",
            "Every page can time out; skip-deep is expensive.",
            "New respond-async job per page; persist jobs (this sample is in-memory).",
            "Per-page wait; still one job only if 202 is returned.",
        ),
        (
            "Excel / Power BI refresh",
            "One long refresh; failure on page N can drop the load.",
            "Refresh waits on Function.InvokeAfter; HTTP timeouts less likely.",
            "Fewer polls than pure async; refresh dialog still blocks.",
        ),
        (
            "Memory",
            "Full concatenated table in the workbook model.",
            "Same once pages are assembled.",
            "Same. 202 does not stream rows while later pages run.",
        ),
        (
            "Recommended pattern",
            "Small pages, cheap filters, $select, timeouts >> page time.",
            "Async first URL if the scan is slow; then measure nextLink.",
            "Async+wait on page 1; short wait or sync on nextLink if cheap.",
        ),
    ]
    for offset, values in enumerate(paging_rows):
        excel_row = 5 + offset
        for col, value in enumerate(values, start=1):
            write_wrapped(paging, excel_row, col, value, ALT_FILL if offset % 2 else None)
        paging.row_dimensions[excel_row].height = 56
    paging["A13"] = "Sequence"
    paging["A13"].font = SUBTITLE
    paging["A14"] = (
        "GET Products?$filter=…   →  (sync | async | async+wait)  →  200 page\n"
        "   if @odata.nextLink exists  →  GET nextLink  →  independent choice of the three modes\n"
        "   repeat until no nextLink."
    )
    paging["A14"].font = CODE
    paging["A14"].alignment = WRAP
    paging.merge_cells("A14:D14")
    paging.row_dimensions[14].height = 52
    paging["A16"] = "Do not"
    paging["A16"].font = SUBTITLE
    paging["A17"] = (
        "Do not assume OData.Feed will send Prefer: respond-async. "
        "Do not rebuild nextLink from $skip unless the service documents conventional paging. "
        "Do not apply a 60s wait to hundreds of cheap continuation pages."
    )
    paging["A17"].font = BODY
    paging["A17"].alignment = WRAP
    paging.merge_cells("A17:D17")
    paging.row_dimensions[17].height = 48
    set_col_widths(paging, {1: 28, 2: 36, 3: 42, 4: 44})
    paging.freeze_panes = "A5"
    paging.sheet_view.showGridLines = False

    add_m_sheet(wb, "M — Sync", ROOT / "power-query" / "Products_Sync.pq")
    add_m_sheet(wb, "M — Async", ROOT / "power-query" / "Products_Async.pq")
    add_m_sheet(wb, "M — Async wait", ROOT / "power-query" / "Products_AsyncWait.pq")

    add_preview_sheet(
        wb,
        "Preview — Sync",
        "Sample of the first rows. The live query follows @odata.nextLink until all 5000 products are loaded (500 per page).",
    )
    add_preview_sheet(
        wb,
        "Preview — Async",
        "Same rows after 202 Accepted and polling /async/{jobId} until AsyncResult 200.",
    )
    add_preview_sheet(
        wb,
        "Preview — Wait",
        "Same rows. If the first page finished within wait=N, Power Query never polled.",
    )

    out = ROOT / "OData-Query-Scenarios.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(path)
